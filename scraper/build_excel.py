"""
Bouwt een Excel-dashboard (data/dashboard.xlsx) op basis van data/latest.json.

Dit is een BONUS-export voor wie liever in Excel werkt. De interactieve
animatie en aan/uit-knoppen per deelnemer zitten in index.html (Excel kan
dat niet); hier krijg je dezelfde cijfers als statische tabellen + lijn-
grafieken, met per deelnemer een vaste, consistente kleur.

Gebruik:
    python build_excel.py
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
INPUT_FILE = DATA_DIR / "latest.json"
OUTPUT_FILE = DATA_DIR / "dashboard.xlsx"

COLOR_PALETTE = [
    "E6194B", "3CB44B", "0082C8", "F58231", "911EB4",
    "46F0F0", "F032E6", "D2A106", "008080", "9A6324",
]

HEADER_FILL = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STAGE_RE = re.compile(r"^Stage (\d+)$")


def load_snapshot() -> dict:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"{INPUT_FILE} bestaat nog niet. Draai eerst scrape.py.")
    return json.loads(INPUT_FILE.read_text(encoding="utf-8"))


def stage_keys_with_data(snapshot: dict) -> list[str]:
    keys = [k for k, v in snapshot["stage_results"].items() if v]
    return sorted(keys, key=lambda k: int(STAGE_RE.match(k).group(1)))


def build_colors_and_names(snapshot: dict):
    slugs = sorted({
        r["participant_slug"] for r in snapshot["classifications"].get("general", [])
        if r.get("participant_slug")
    })
    colors = {slug: COLOR_PALETTE[i % len(COLOR_PALETTE)] for i, slug in enumerate(slugs)}
    names = {
        r["participant_slug"]: r.get("Participant", r["participant_slug"])
        for r in snapshot["classifications"].get("general", [])
        if r.get("participant_slug")
    }
    return colors, names


def style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def sheet_overzicht(wb: Workbook, snapshot: dict) -> None:
    ws = wb.active
    ws.title = "Overzicht"
    ws["A1"] = "Algemeen klassement (belangrijkste stand)"
    ws["A1"].font = Font(bold=True, size=14)

    row_cursor = 3
    for tab_id, label in [
        ("general", "* Algemeen klassement"),
        ("points", "Puntenklassement"),
        ("mountain", "Bergklassement"),
        ("youth", "Jongerenklassement"),
    ]:
        rows = snapshot["classifications"].get(tab_id, [])
        if not rows:
            continue
        ws.cell(row=row_cursor, column=1, value=label).font = Font(bold=True, size=12)
        row_cursor += 1
        headers = [h for h in rows[0].keys() if h != "participant_slug"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row_cursor, column=c, value=h)
        style_header_row(ws, row_cursor, len(headers))
        row_cursor += 1
        for r in rows:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row_cursor, column=c, value=r.get(h, ""))
            row_cursor += 1
        row_cursor += 2

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_matrix(snapshot: dict, value_key: str, colors: dict, names: dict):
    stage_keys = stage_keys_with_data(snapshot)
    slugs = sorted(colors.keys(), key=lambda s: names.get(s, s))
    matrix = {slug: [] for slug in slugs}
    for stage_key in stage_keys:
        rows = {r.get("participant_slug"): r for r in snapshot["stage_results"].get(stage_key, [])}
        for slug in slugs:
            row = rows.get(slug)
            val = None
            if row:
                try:
                    val = float(row.get(value_key, "-"))
                except (TypeError, ValueError):
                    val = None
            matrix[slug].append(val)
    return stage_keys, slugs, matrix


def sheet_verloop(wb: Workbook, snapshot: dict, value_key: str, sheet_name: str,
                   colors: dict, names: dict, prominent: bool = False) -> None:
    stage_keys, slugs, matrix = build_matrix(snapshot, value_key, colors, names)
    if not stage_keys:
        return

    ws = wb.create_sheet(("STAR " if prominent else "") + sheet_name)
    ws.cell(row=1, column=1, value="Etappe")
    for c, slug in enumerate(slugs, start=2):
        ws.cell(row=1, column=c, value=names.get(slug, slug))
    style_header_row(ws, 1, len(slugs) + 1)

    for r, stage_key in enumerate(stage_keys, start=2):
        ws.cell(row=r, column=1, value=stage_key)
        for c, slug in enumerate(slugs, start=2):
            ws.cell(row=r, column=c, value=matrix[slug][r - 2])

    n_rows, n_cols = len(stage_keys), len(slugs)

    chart = LineChart()
    chart.title = sheet_name
    chart.style = 2
    chart.y_axis.title = value_key
    chart.x_axis.title = "Etappe"
    chart.width = 30
    chart.height = 15 if prominent else 12

    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
    for i, slug in enumerate(slugs):
        col = i + 2
        # min_row=2: alleen de data, GEEN header-rij. De naam zetten we
        # hieronder direct en statisch (s.tx), zodat 'm meteen de juiste
        # deelnemersnaam toont in Excel i.p.v. "Series1"/"Series2" (dat
        # laatste gebeurt als je title_from_data gebruikt met een
        # formule-verwijzing die pas na herberekening zichtbaar wordt).
        data_ref = Reference(ws, min_col=col, min_row=2, max_row=1 + n_rows)
        series = Series(data_ref)
        series.tx = SeriesLabel(v=names.get(slug, slug))
        series.marker = Marker(symbol="circle", size=5)
        series.graphicalProperties = GraphicalProperties()
        series.graphicalProperties.line = LineProperties(solidFill=colors[slug], w=24000)
        chart.series.append(series)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(n_cols + 3)}2")

    for col in range(1, n_cols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 16


def sheet_stijgers(wb: Workbook, snapshot: dict, colors: dict, names: dict) -> None:
    ws = wb.create_sheet("Stijgers")
    ws["A1"] = "Grootste stijgers / dalers (algemeen klassement, totaalpunten)"
    ws["A1"].font = Font(bold=True, size=14)

    stage_keys, slugs, matrix = build_matrix(snapshot, "Pts", colors, names)
    if len(stage_keys) < 2:
        ws["A3"] = "Nog niet genoeg etappes met data (minimaal 2 nodig)."
        return

    def diff_table(n_back: int, label: str, start_row: int) -> int:
        ws.cell(row=start_row, column=1, value=label).font = Font(bold=True)
        headers = ["Deelnemer", "Punten toen", "Punten nu", "Verschil"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=start_row + 1, column=c, value=h)
        style_header_row(ws, start_row + 1, len(headers))

        idx_now = len(stage_keys) - 1
        idx_then = max(0, idx_now - n_back)
        rows = []
        for slug in slugs:
            now_val, then_val = matrix[slug][idx_now], matrix[slug][idx_then]
            if now_val is None or then_val is None:
                continue
            rows.append((names.get(slug, slug), then_val, now_val, now_val - then_val))
        rows.sort(key=lambda r: r[3], reverse=True)

        r = start_row + 2
        for name, then_val, now_val, delta in rows:
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=then_val)
            ws.cell(row=r, column=3, value=now_val)
            ws.cell(row=r, column=4, value=delta)
            r += 1
        return r + 2

    next_row = diff_table(1, "Laatste etappe", 3)
    diff_table(min(7, len(stage_keys) - 1), "Laatste 7 etappes", next_row)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22


def sheet_bijzaak(wb: Workbook, snapshot: dict) -> None:
    ws = wb.create_sheet("Bijzaak - renners en historie")
    row = 1
    ws.cell(row=row, column=1, value="Most picked (renners)").font = Font(bold=True, size=13)
    row += 1
    if snapshot.get("most_picked"):
        headers = list(snapshot["most_picked"][0].keys())
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1
        for r in snapshot["most_picked"]:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row, column=c, value=r.get(h, ""))
            row += 1
    row += 2

    ws.cell(row=row, column=1, value="Most points (renners)").font = Font(bold=True, size=13)
    row += 1
    if snapshot.get("most_points"):
        headers = list(snapshot["most_points"][0].keys())
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1
        for r in snapshot["most_points"]:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row, column=c, value=r.get(h, ""))
            row += 1
    row += 2

    ws.cell(row=row, column=1, value="History (eerdere jaren)").font = Font(bold=True, size=13)
    row += 1
    if snapshot.get("history"):
        headers = list(snapshot["history"][0].keys())
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1
        for r in snapshot["history"]:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row, column=c, value=r.get(h, ""))
            row += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main() -> None:
    snapshot = load_snapshot()
    colors, names = build_colors_and_names(snapshot)

    wb = Workbook()
    sheet_overzicht(wb, snapshot)
    sheet_verloop(wb, snapshot, "Pts", "Verloop Algemeen (Pts)", colors, names, prominent=True)
    sheet_verloop(wb, snapshot, "GC", "Verloop GC-punten", colors, names)
    sheet_verloop(wb, snapshot, "PC", "Verloop Puntenklassement", colors, names)
    sheet_verloop(wb, snapshot, "MC", "Verloop Bergklassement", colors, names)
    sheet_stijgers(wb, snapshot, colors, names)
    sheet_bijzaak(wb, snapshot)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Dashboard opgeslagen: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
